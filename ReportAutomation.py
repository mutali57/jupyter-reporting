# -*- coding: utf-8 -*-
"""
Created on Fri Jul 15 12:10:22 2022

@author: mutali
"""
import openpyxl as xl
from openpyxl.chart import LineChart, Reference

import win32com.client
import PIL
from PIL import ImageGrab, Image
import os
import sys

from docx.shared import Cm
from docxtpl import DocxTemplate, InlineImage
from docx.shared import Cm, Inches, Mm, Emu
import random
import datetime
import matplotlib.pyplot as plt


######## Generate automated excel workbook ########

workbook = xl.load_workbook('Moshesh Perfomance.xlsm')
sheet_1 = workbook['Sheet1']
  

######## Extract chart image from Excel workbook ########

input_file = "Book1.xlsx"

output_image = "Chart.png"
operation = win32com.client.Dispatch("Excel.Application")
operation.Visible = 0
operation.DisplayAlerts = 0
workbook_2 = operation.Workbooks.Open(input_file)
sheet_2 = operation.Sheets(1)

for x, chart in enumerate(sheet_2.Shapes):
    chart.Copy()
    image = ImageGrab.grabclipboard()
    image.save('', 'png')
    pass
workbook_2.Close(True)
operation.Quit()
"""
######## Generating automated word document ########

template = DocxTemplate('template.docx')

#Generate list of random values
table_contents = []

for i in range(2, sheet_1.max_row + 1):
    
    table_contents.append({
        'Index': i-1,
        'Power': sheet_1.cell(i, 1).value,
        'Current': sheet_1.cell(i, 2).value,
        'Voltage': sheet_1.cell(i, 3).value
        })

#Import saved figure
image = InlineImage(template,'party_banner_0.png',Cm(10))

#Declare template variables
context = {
    'title': 'Automated Report',
    'day': datetime.datetime.now().strftime('%d'),
    'month': datetime.datetime.now().strftime('%b'),
    'year': datetime.datetime.now().strftime('%Y'),
    'table_contents': table_contents,
    'image': image
    }

#Render automated report
template.render(context)
template.save('Automated_report.docx')"""